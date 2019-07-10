# Based almost entirely on https://bitbucket.org/adimian/xlsx-diff/src/default/
# License: MIT License (MIT/Expat)
# Author: Eric Gazoni (eric.gazoni@adimian.com)

import openpyxl


try:
    # Python 3
    from itertools import zip_longest
except ImportError:
    from itertools import izip_longest as zip_longest


class Difference(object):
    def __init__(self, **kwargs):
        for k, v in kwargs.items():
            setattr(self, k, v)
            self.__dict__[k] = v

    def __repr__(self):
        return u'%s: {%s}' % (self.__class__.__name__,
                              u', '.join([k + ': ' + repr(self.__dict__[k])
                                          for k in self.__slots__]))


class MissingWorksheet(Difference):
    __slots__ = ('worksheet', 'missing_in')


class MissingCell(Difference):
    __slots__ = ('worksheet', 'coordinate', 'missing_in')


class CellDifference(Difference):
    __slots__ = ('worksheet', 'coordinate', 'kind', 'expected', 'found')


def diff(ref, other, ignores=(), precision=None, typeless=False):
    ''' compare a reference workbook with another workbook
    @param ignores: list of ignore classes to ignore
    @param precision: number of decimal digits to keep when comparing floats
    @param typeless: whether we consider number-looking strings as number or not

    @return: (workbook-level differences, dict with per-sheet differences)
    '''

    kwargs = dict(read_only=True, keep_vba=False, data_only=True)

    if precision is not None:
        if precision < 0:
            raise ValueError('negative precision is meaningless')
        else:
            precision = 10 ** (-precision)

    wb1 = openpyxl.load_workbook(ref, **kwargs)
    wb2 = openpyxl.load_workbook(other, **kwargs)

    wb1_sheets = set(wb1.sheetnames)
    wb2_sheets = set(wb2.sheetnames)

    wb_differences = []
    if not MissingWorksheet in ignores:
        for sh in (wb1_sheets - wb2_sheets):
            wb_differences.append(MissingWorksheet(worksheet=sh,
                                                   missing_in='reference'))

        for sh in (wb2_sheets - wb1_sheets):
            wb_differences.append(MissingWorksheet(worksheet=sh,
                                                   missing_in='other'))

    sheet_differences = {}
    for sh in (wb1_sheets & wb2_sheets):
        changes = sheet_changes(wb1, wb2, sh, ignores, precision, typeless)
        if changes:
            sheet_differences[sh] = changes

    return wb_differences, sheet_differences


def sheet_changes(wb1, wb2, sheet_name, ignores, precision, typeless):
    sheet1 = wb1[sheet_name]
    sheet2 = wb2[sheet_name]

    diffs = []
    miss = MissingCell not in ignores
    diff = CellDifference not in ignores
    for r1, r2 in zip_longest(sheet1.rows, sheet2.rows):
        if r1 is None:
            if miss:
                for cell in r2:
                    if not empty(cell):
                        diffs.append(MissingCell(worksheet=sheet_name,
                                                 coordinate=cell.coordinate,
                                                 missing_in='reference'))
        elif r2 is None:
            if miss:
                for cell in r1:
                    if not empty(cell):
                        diffs.append(MissingCell(worksheet=sheet_name,
                                                 coordinate=cell.coordinate,
                                                 missing_in='other'))
        else:
            for c1, c2 in zip_longest(r1, r2):
                if empty(c1) and empty(c2):
                    continue
                elif empty(c1) and not empty(c2):
                    if miss:
                        diffs.append(MissingCell(worksheet=sheet_name,
                                                 coordinate=c2.coordinate,
                                                 missing_in='reference'))
                elif empty(c2) and not empty(c1):
                    if miss:
                        diffs.append(MissingCell(worksheet=sheet_name,
                                                 coordinate=c1.coordinate,
                                                 missing_in='other'))
                elif diff:
                    if c1.value != c2.value:
                        v1 = c1.value
                        v2 = c2.value
                        if typeless:
                            try:
                                v1, v2 = float(v1), float(v2)
                            except:
                                pass
                        if isinstance(v1, float) and isinstance(v2, float):
                            if precision is not None:
                                if abs(v1 - v2) < precision:
                                    continue
                            else:
                                if v1 == v2:
                                    continue
                        diffs.append(CellDifference(worksheet=sheet_name,
                                                    coordinate=c1.coordinate,
                                                    kind='value',
                                                    expected=v1,
                                                    found=v2))
    diffs.sort(key=klass)
    return diffs


def empty(cell):
    if cell and not cell.value is None:
        return False
    return True


def klass(x):
    return x.__class__.__name__
